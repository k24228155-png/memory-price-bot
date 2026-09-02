import pandas as pd
import time
import os
import re
import requests
import sys
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from datetime import datetime
import matplotlib.pyplot as plt

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

EXCEL_FILE = "Memory_Daily_Report.xlsx"

TARGETS = {
    "DRAM_Spot": "https://www.trendforce.com.tw/price/dram/dram_spot",
    "Flash_Spot": "https://www.trendforce.com.tw/price/flash/flash_spot",
    "Wafer_Spot": "https://www.trendforce.com.tw/price/flash/wafer_spot",
    "DRAM_Contract": "https://www.trendforce.com.tw/price/dram/dram_contract",
    "Flash_Contract": "https://www.trendforce.com.tw/price/flash/flash_contract"
}

def clean_change_value(text):
    text = text.replace('▲', '').replace('▼', '').replace('%', '').strip()
    try:
        return float(text)
    except ValueError:
        return 0.0

def format_header_name(name, type_str):
    name = re.sub(r'\s*\d{4}/\d{4}|\s*\d{4}', '', name).strip()
    name = name.replace(' (', '\n(')
    return f"{name}\n{type_str}"

def get_page_data(driver, url, sheet_name):
    print(f"\n⏳ 正在抓取 [{sheet_name}] ...")
    try:
        driver.get(url)
        time.sleep(6) 
        
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        table = soup.find('table')
        if not table: return None
            
        today_str = datetime.now().strftime("%Y-%m-%d")
        new_row = {"Date": today_str}
        
        headers = table.find_all('th')
        avg_idx, change_idx = 5, 6
        for i, th in enumerate(headers):
            text = th.text.lower()
            if 'avg' in text or 'average' in text: avg_idx = i
            if 'change' in text or '%' in text: change_idx = i
            
        rows = table.find_all('tr')
        parsed_count = 0
        for tr in rows:
            tds = tr.find_all('td')
            if len(tds) > max(avg_idx, change_idx):
                item_name = tds[0].text.strip().replace('\n', '')
                if item_name and "Item" not in item_name:
                    try:
                        avg_val = float(tds[avg_idx].text.strip())
                        change_val = clean_change_value(tds[change_idx].text.strip())
                        col_avg = format_header_name(item_name, "均價")
                        col_change = format_header_name(item_name, "漲跌幅")
                        new_row[col_avg] = avg_val
                        new_row[col_change] = change_val
                        parsed_count += 1
                    except Exception:
                        pass
        if parsed_count > 0:
            print(f"✅ [{sheet_name}] 成功抓取 {parsed_count} 筆資料！")
            return pd.DataFrame([new_row])
        return None
    except Exception:
        return None

def draw_trend_chart(df, sheet_name):
    if len(df) < 2: return 
    plt.figure(figsize=(14, 7))
    avg_cols = [col for col in df.columns if '均價' in str(col)]
    if not avg_cols: return
        
    for col in avg_cols:
        display_name = str(col).replace('\n均價', '').replace('\n', ' ') 
        plt.plot(df['Date'], df[col], marker='o', linewidth=2, label=display_name)
    
    plt.rcParams['font.sans-serif'] = ['Microsoft JhengHei'] 
    plt.rcParams['axes.unicode_minus'] = False
    plt.title(f'{sheet_name} - 均價趨勢圖', fontsize=14, fontweight='bold')
    plt.xlabel('日期')
    plt.ylabel('平均價格 (USD)')
    plt.xticks(rotation=45)
    plt.grid(True, linestyle='--', alpha=0.5)
    plt.legend(loc='center left', bbox_to_anchor=(1, 0.5), fontsize=9)
    plt.tight_layout()
    plt.savefig(f"{sheet_name}_Chart.png")
    plt.close()

def format_excel_style(filename):
    if not os.path.exists(filename): return
    wb = load_workbook(filename)
    date_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")   
    avg_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")    
    change_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") 
    header_font = Font(bold=True)
    center_wrap_align = Alignment(horizontal="center", vertical="center", wrap_text=True) 

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.row_dimensions[1].height = 60
        for col in ws.columns:
            col_letter = col[0].column_letter 
            ws.column_dimensions[col_letter].width = 13 if col_letter == 'A' else 11.5 
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = center_wrap_align
                if cell.row == 1:
                    cell.font = header_font
                    if cell.value == "Date": cell.fill = date_fill
                    elif cell.value and "均價" in str(cell.value): cell.fill = avg_fill
                    elif cell.value and "漲跌幅" in str(cell.value): cell.fill = change_fill
    wb.save(filename)

def update_excel_and_draw_charts(all_new_data):
    all_history = {}
    if os.path.exists(EXCEL_FILE):
        try: all_history = pd.read_excel(EXCEL_FILE, sheet_name=None)
        except Exception: pass
            
    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
        for sheet_name, new_df in all_new_data.items():
            if sheet_name in all_history:
                old_cols = set(all_history[sheet_name].columns)
                new_cols = set(new_df.columns)
                if len(old_cols.intersection(new_cols)) < 2:
                    combined_df = new_df
                else:
                    combined_df = pd.concat([all_history[sheet_name], new_df], ignore_index=True)
                    combined_df.drop_duplicates(subset=['Date'], keep='last', inplace=True)
                    combined_df = combined_df[new_df.columns] 
            else:
                combined_df = new_df
            combined_df.to_excel(writer, sheet_name=sheet_name, index=False)
            draw_trend_chart(combined_df, sheet_name)
    format_excel_style(EXCEL_FILE)

def send_line_messaging_api():
    """🌟 讀取 GitHub Secrets 並透過 LINE Messaging API 發送文字與圖片"""
    token = os.environ.get("LINE_TOKEN")
    user_id = os.environ.get("LINE_USER_ID")
    repo = os.environ.get("GITHUB_REPOSITORY") 
    
    if not token or not user_id:
        print("💡 未偵測到 LINE 密鑰 (Token 或 User ID)，無法發送通知。")
        return
        
    url = "https://api.line.me/v2/bot/message/push"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    today_str = datetime.now().strftime("%Y-%m-%d")
    text_msg = f"📊 【記憶體報價自動通知】\n📅 日期：{today_str}\n✅ 雲端 5 大網頁報價已抓取完畢，趨勢圖如附！"
    
    # 發送文字
    requests.post(url, headers=headers, json={"to": user_id, "messages": [{"type": "text", "text": text_msg}]})
    
    # 發送圖片
    image_messages = []
    base_url = f"https://raw.githubusercontent.com/{repo}/main/"
    for sheet_name in TARGETS.keys():
        chart_file = f"{sheet_name}_Chart.png"
        if os.path.exists(chart_file):
            img_url = f"{base_url}{chart_file}?t={int(time.time())}"
            image_messages.append({
                "type": "image",
                "originalContentUrl": img_url,
                "previewImageUrl": img_url
            })
            
    if image_messages:
        r = requests.post(url, headers=headers, json={"to": user_id, "messages": image_messages})
        if r.status_code == 200:
            print("🚀 LINE 通知與圖表已成功發送！")
        else:
            print(f"❌ 圖片發送失敗，錯誤碼: {r.status_code}, {r.text}")

def main():
    chrome_options = Options()
    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("window-size=1920,1080")
    chrome_options.add_argument("user-agent=Mozilla/5.0")

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
    all_new_data = {}
    for sheet_name, url in TARGETS.items():
        df = get_page_data(driver, url, sheet_name)
        if df is not None: all_new_data[sheet_name] = df
    driver.quit()
    
    if all_new_data:
        update_excel_and_draw_charts(all_new_data)
    else:
        print("\n❌ 今日無有效資料。")

if __name__ == "__main__":
    # 🌟 透過系統參數決定要執行爬蟲還是發送通知
    if len(sys.argv) > 1 and sys.argv[1] == "--notify":
        print("🔔 爬蟲與存檔已完成，開始執行 LINE 通知發送程序...")
        send_line_messaging_api()
    else:
        main()
